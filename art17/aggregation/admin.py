# coding=utf-8
from collections import OrderedDict
from copy import deepcopy
from string import upper

import flask
from flask import redirect, url_for, render_template, request, flash, Response
from flask import jsonify
from wtforms import Form, TextField, SelectField
from flask.ext.principal import Permission
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from sqlalchemy import or_

from art17 import dal, models, ROLE_FINAL
from art17.aggregation.checklist import create_checklist
from art17.aggregation.forms import CompareForm, PreviewForm, RefValuesForm
from art17.aggregation.utils import (
    get_checklist, get_reporting_id, get_species_checklist,
    get_habitat_checklist, valid_checklist, sum_of_reports,
)
from art17.auth import require, need
from art17.common import get_datasets, TemplateView, get_year_end
from art17.models import (
    Dataset,
    db,
    DATASET_STATUSES_DICT,
    DATASET_STATUSES_LIST,
    STATUS_NEW,
    STATUS_CONSULTATION,
    LuGrupSpecie,
    LuHdSpecies,
    LuHabitattypeCodes,
)
from art17.aggregation import (
    aggregation,
    perm_edit_refvals,
    perm_view_refvals,
    MIMETYPE)
from art17.aggregation.refvalues import (
    load_species_refval,
    refvalue_ok,
    load_habitat_refval,
    get_subject_refvals, get_subject_refvals_wip, set_subject_refvals_wip,
    get_subject_refvals_mixed,
    save_species_refval, save_habitat_refval,
)

REGIONS = {
    'ALP': 'Alpina',
    'CON': 'Continentala',
    'PAN': 'Panonica',
    'STE': 'Stepica',
    'BLS': 'Boreala',
    'MBLS': 'Marea Neagra',
}

REFGROUPS = {
    'range': 'Areal',
    'magnitude': 'Magnitudine Areal',
    'population_units': 'Unitati populatie',
    'population_magnitude': 'Magnitudine Populatie',
    'population_range': 'Populatie',
    'coverage_range': 'Suprafata',
    'coverage_magnitude': 'Magnitudine suprafata',
    'typical_species': 'Specii tipice',
    'threats': 'Amenintari',
    'pressures': 'Presiuni',
    'habitat': 'Habitat',
}

EXTRA_COLUMS = ['COD', 'NUME', 'BIOREGIUNE']


def get_struct(refvals):
    struct = refvals.values()[0]
    return {REFGROUPS.get(k, k.capitalize()): EXTRA_COLUMS + map(upper, d.keys())
            for k, d in struct.iteritems()}


def reverse(struct):
    reverse_refgroups = {v: k for k, v in REFGROUPS.iteritems()}
    rev = {reverse_refgroups[group]:
           {f.capitalize(): '' for f in fields if f not in EXTRA_COLUMS}
           for group, fields in struct.iteritems()}
    return rev, reverse_refgroups


def get_refvals(page):
    if page == 'habitat':
        return load_habitat_refval()
    elif page == 'species':
        return load_species_refval()
    else:
        raise NotImplementedError()


def get_checklists():
    return Dataset.query.filter_by(checklist=True).order_by(
        Dataset.date.desc())


def parse_checklist(checklist_qs):
    result = OrderedDict()
    for item in checklist_qs:
        key = (item.code, item.name)
        if key not in result:
            result[key] = {'info': item, 'regions': [item.bio_region]}
        else:
            result[key]['regions'].append(item.bio_region)
    return result


@aggregation.app_template_global('can_delete_dataset')
def can_delete_dataset(dataset):
    if dataset.status == STATUS_NEW:
        return True
    return False


@aggregation.route('/admin/')
@require(Permission(need.admin, need.reporter))
def admin():
    return redirect(url_for('.checklists'))


@aggregation.route('/admin/checklists/')
@require(Permission(need.admin, need.reporter))
def checklists():
    checklists = get_checklists()

    default_checklist = {
        'species_checklist': get_species_checklist().all(),
        'habitat_checklist': get_habitat_checklist().all(),
    }
    years = (db.session.query(Dataset.year_start)
             .filter(or_(Dataset.preview == False, Dataset.preview == None))
             .order_by(Dataset.year_start.desc())
             .distinct())
    periods = [(year, get_year_end(year)) for year, in years]

    return render_template(
        'aggregation/admin/checklists.html',
        page='checklist',
        default_list=default_checklist,
        checklists=checklists,
        periods=periods,
    )


@aggregation.route('/admin/checklist/initial/')
@aggregation.route('/admin/checklist/<dataset_id>/')
@require(Permission(need.admin, need.reporter))
def checklist(dataset_id=None):
    species = get_species_checklist(dataset_id=dataset_id)
    habitats = get_habitat_checklist(dataset_id=dataset_id)
    species_dict = parse_checklist(species)
    habitats_dict = parse_checklist(habitats)

    return render_template(
        'aggregation/admin/checklist.html',
        species_dict=species_dict,
        habitats_dict=habitats_dict,
        page='checklist',
    )


@aggregation.route('/admin/checklist/create/', methods=('GET', 'POST'))
@aggregation.route('/admin/checklist/create/<save_current>',
                   methods=('GET', 'POST'))
@require(Permission(need.admin, need.reporter))
def create(save_current=False):
    if request.method == "POST":
        create_checklist(save_current)
        return redirect(url_for('.checklists'))
    return render_template(
        'aggregation/admin/checklist_create.html',
        page='checklist',
    )


class ChecklistForm(Form):
    comment = TextField()


class DatasetForm(Form):
    comment = TextField()
    status = SelectField(choices=DATASET_STATUSES_LIST)

    def __init__(self, *args, **kwargs):
        self.obj = kwargs['obj']
        return super(DatasetForm, self).__init__(*args, **kwargs)

    def validate_status(self, field):
        if int(field.data) != STATUS_CONSULTATION:
            return True
        active_exists = Dataset.query.filter(
            Dataset.id != self.obj.id,
            Dataset.year_start == self.obj.year_start,
            Dataset.status == STATUS_CONSULTATION,
        ).count()
        if active_exists:
            field.errors.append(
                'Există deja un set de date activ. Nu poate exista mai mult de'
                ' un set de date activ la un moment dat.')
            return False
        return True


@aggregation.route('/admin/checklist/<dataset_id>/edit/',
                   methods=('GET', 'POST'))
@require(Permission(need.admin, need.reporter))
def edit_checklist(dataset_id):
    dataset = get_checklists().filter_by(id=dataset_id).first()

    if request.method == 'POST':
        form = ChecklistForm(request.form, obj=dataset)
        if form.validate():
            form.populate_obj(dataset)
            db.session.commit()
            return redirect(url_for('.admin'))
    else:
        form = ChecklistForm(obj=dataset)

    return render_template(
        'aggregation/admin/edit_checklist.html',
        page='checklist',
        dataset=dataset,
        form=form,
    )


@aggregation.route('/admin/dataset/<dataset_id>/edit/',
                   methods=('GET', 'POST'))
@require(Permission(need.admin, need.reporter))
def edit_dataset(dataset_id):
    dataset = get_datasets().filter_by(id=dataset_id).first()

    if request.method == 'POST':
        form = DatasetForm(request.form, obj=dataset)
        if form.validate():
            form.populate_obj(dataset)
            db.session.commit()
            flash("Form successfully updated", 'success')
            return jsonify({'status': 'success',
                            'url': url_for('.admin')})
        else:
            return jsonify({'status': 'error',
                            'errors': form.errors})
    else:
        form = DatasetForm(obj=dataset)

    return render_template(
        'aggregation/admin/edit_dataset.html',
        page='checklist',
        dataset=dataset,
        form=form,
    )


class ReferenceValues(TemplateView):
    template_name = 'aggregation/admin/reference_values.html'
    decorators = [require(Permission(need.admin, need.reporter))]

    def get_context(self, **kwargs):
        checklist_id = get_reporting_id()
        current_checklist = get_checklist(checklist_id)
        checklist_id = current_checklist.id

        species_refvals = load_species_refval()
        species_checklist = get_species_checklist(dataset_id=checklist_id)
        species_data = parse_checklist_ref(species_checklist)

        species_list = get_species_checklist(groupped=True,
                                             dataset_id=checklist_id)

        habitat_refvals = load_habitat_refval()
        habitat_checklist = get_habitat_checklist(dataset_id=checklist_id)
        habitat_data = parse_checklist_ref(habitat_checklist)
        habitat_list = get_habitat_checklist(distinct=True,
                                             dataset_id=checklist_id,
                                             groupped=True)
        relevant_regions = (
            {s.bio_region for s in species_checklist}.union(
                {h.bio_region for h in habitat_checklist}
            ))
        bioreg_list = dal.get_biogeo_region_list(relevant_regions)

        groups = dict(
            LuGrupSpecie.query
            .with_entities(LuGrupSpecie.code, LuGrupSpecie.description)
        )

        return dict(
            species_refvals=species_refvals,
            species_data=species_data,
            species_list=species_list,
            habitat_refvals=habitat_refvals,
            habitat_data=habitat_data,
            habitat_list=habitat_list,
            bioreg_list=bioreg_list,
            GROUPS=groups,
            current_checklist=current_checklist,
            page='refvalues',
        )


aggregation.add_url_rule('/admin/reference_values',
                         view_func=ReferenceValues.as_view('reference_values'))


def process_xls(wb, struct):
    d = {}
    reversed_struct, reversed_refgroups = reverse(struct)
    for sheet in wb.get_sheet_names():
        ws = wb[sheet]
        col_names = [c.value.capitalize() for c in ws.rows[0] if c.value]
        for row in ws.rows[1:]:
            codereg = row[0].value + '-' + row[2].value
            if codereg not in d:
                d[codereg] = deepcopy(reversed_struct)
            for col in range(3, len(col_names)):
                d[codereg][reversed_refgroups[sheet]][col_names[col]] = \
                    str(row[col].value or '')
    return d


class ReferenceValuesUpdate(TemplateView):
    template_name = 'aggregation/admin/refvals_update.html'
    decorators = [require(Permission(need.admin))]

    def get_context(self, **kwargs):
        subject = kwargs.pop('subject')
        form = RefValuesForm(request.files)
        return dict(
            form=form,
            subject=subject,
            page='refvalues',
        )

    def post(self, **kwargs):
        context = self.get_context(**kwargs)
        form = context['form']
        subject = context['subject']
        if form.validate():
            required_struct = get_struct(get_refvals(context['subject']))
            wb = load_workbook(form.excel_doc.data)
            struct = {sheet: [r.value for r in wb[sheet].rows[0] if r.value]
                      for sheet in wb.get_sheet_names()}
            if required_struct == struct:
                d = process_xls(wb, struct)
                if subject == 'species':
                    save_species_refval(d)
                elif subject == 'habitat':
                    save_habitat_refval(d)
                else:
                    raise NotImplementedError
                flash(u'Noile valori de referință au fost salvate.', 'success')
            else:
                flash(u'Documentul Excel nu este în formatul acceptat.',
                      'danger')
        return render_template(self.template_name, **context)

aggregation.add_url_rule(
    '/admin/reference_values/<string:subject>/update',
    view_func=ReferenceValuesUpdate.as_view('refvals_update'))


@aggregation.app_context_processor
def inject_globals():
    return {
        'checklists': get_checklists(),
        'datasets': get_datasets(),
        'DATASET_STATUSES': DATASET_STATUSES_DICT,
        'refvalue_ok': refvalue_ok,
        'sum_of_reports': sum_of_reports,
        'REFGROUPS': REFGROUPS,
    }


@aggregation.route('/admin/compare/select', methods=('GET', 'POST'))
@require(Permission(need.admin, need.reporter))
def compare():
    if request.method == 'POST':
        form = CompareForm(request.form)

        if form.validate():
            return flask.redirect(
                flask.url_for(
                    '.compare_datasets',
                    dataset1=form.dataset1.data,
                    dataset2=form.dataset2.data)
            )
    else:
        form = CompareForm()
    return flask.render_template('aggregation/compare.html', form=form,
                                 page='compare')


@aggregation.route('/admin/compare/<int:dataset1>/<int:dataset2>/')
@require(Permission(need.admin, need.reporter))
def compare_datasets(dataset1, dataset2):
    d1 = models.Dataset.query.get_or_404(dataset1)
    d2 = models.Dataset.query.get_or_404(dataset2)

    conclusions_s_d1 = d1.species_objs.filter_by(cons_role=ROLE_FINAL)
    conclusions_s_d2 = d2.species_objs.filter_by(cons_role=ROLE_FINAL)

    relevant_regions = set([r[0] for r in (
        list(conclusions_s_d1.with_entities(models.DataSpeciesRegion.region)
             .distinct()) +
        list(conclusions_s_d2.with_entities(models.DataSpeciesRegion.region)
             .distinct())
    ) if r[0]])

    s_data = {}
    for r in conclusions_s_d1:
        if r.species not in s_data:
            s_data[r.species] = {'d1': {}, 'd2': {}}
        s_data[r.species]['d1'][r.region] = r
    for r in conclusions_s_d2:
        if r.species not in s_data:
            s_data[r.species] = {'d1': {}, 'd2': {}}
        s_data[r.species]['d2'][r.region] = r
    if None in s_data:
        del s_data[None]

    s_stat = {'objs': 0, 'diff': 0}
    for k, v in s_data.iteritems():
        for reg, ass in v['d1'].iteritems():
            ass2 = v['d2'].get(reg, None)
            if (not ass2 or
                    ass2.conclusion_assessment != ass.conclusion_assessment):
                s_stat['diff'] += 1
            s_stat['objs'] += 1

    conclusions_h_d1 = d1.habitat_objs.filter_by(cons_role=ROLE_FINAL)
    conclusions_h_d2 = d2.habitat_objs.filter_by(cons_role=ROLE_FINAL)

    h_data = {}
    for r in conclusions_h_d1:
        if r.habitat not in h_data:
            h_data[r.habitat] = {'d1': {}, 'd2': {}}
        h_data[r.habitat]['d1'][r.region] = r
    for r in conclusions_h_d2:
        if r.habitat not in h_data:
            h_data[r.habitat] = {'d1': {}, 'd2': {}}
        h_data[r.habitat]['d2'][r.region] = r
    if None in h_data:
        del h_data[None]

    h_stat = {'objs': 0, 'diff': 0}
    for k, v in h_data.iteritems():
        for reg, ass in v['d1'].iteritems():
            ass2 = v['d2'].get(reg, None)
            if (not ass2 or
                    ass2.conclusion_assessment != ass.conclusion_assessment):
                h_stat['diff'] += 1
            h_stat['objs'] += 1

    bioreg_list = dal.get_biogeo_region_list(relevant_regions)
    return render_template(
        'aggregation/compare_datasets.html',
        species_data=s_data, dataset1=d1, dataset2=d2, bioreg_list=bioreg_list,
        habitat_data=h_data, s_stat=s_stat, h_stat=h_stat,
        page='compare',
    )


@aggregation.route('/manage/reference_values/', methods=['GET', 'POST'])
@aggregation.route('/manage/reference_values/<page>', methods=['GET', 'POST'])
@require(perm_edit_refvals())
def manage_refvals(page='habitat'):
    current_checklist = valid_checklist()
    checklist_id = current_checklist.id

    dataset = None
    form = PreviewForm(formdata=request.form, page=page,
                       checklist_id=checklist_id)
    report = None
    if request.method == 'POST':
        if form.validate():
            subject = form.subject.data
            return redirect(
                url_for('.manage_refvals_form', page=page, subject=subject))
    return flask.render_template(
        'aggregation/manage/refvals_start.html',
        **{
            'form': form, 'dataset': dataset, 'report': report, 'page': page,
            'current_checklist': current_checklist,
            'endpoint': '.manage_refvals',
        })


class ManageReferenceValues(ReferenceValues):
    template_name = 'aggregation/manage/reference_values.html'
    decorators = [require(perm_view_refvals())]


aggregation.add_url_rule('/manage/reference_values/table',
                         view_func=ManageReferenceValues.as_view(
                             'manage_refvals_table'))


@aggregation.route('/manage/reference_values/<page>/form/<subject>',
                   methods=['GET', 'POST'])
@require(perm_edit_refvals())
def manage_refvals_form(page, subject):
    data = get_subject_refvals(page, subject)

    if request.method == "POST":
        set_subject_refvals_wip(page, subject, request.form)
        flask.flash(u"Valori actualizate", 'success')

    extra = get_subject_refvals_wip(page, subject)
    full = get_subject_refvals_mixed(page, subject)
    checklist_id = get_reporting_id()
    current_checklist = get_checklist(checklist_id)
    checklist_id = current_checklist.id
    if page == 'habitat':
        names = get_habitat_checklist(distinct=True, dataset_id=checklist_id)
    elif page == 'species':
        names = get_species_checklist(distinct=True, dataset_id=checklist_id)
    name_and_code = names.filter_by(code=subject).first()[1]
    _, _, name = name_and_code.partition(subject)

    return flask.render_template(
        'aggregation/manage/refvals_form.html', page=page, subject=subject,
        data=data, extra=extra, full=full, name=name, REGIONS=REGIONS,
    )


@aggregation.route('/manage/reference_values/<page>/form/<subject>/download',
                   methods=['GET'])
def download_refvals(page, subject):
    checklist_id = get_reporting_id()
    current_checklist = get_checklist(checklist_id)
    checklist_id = current_checklist.id
    if page == 'habitat':
        names = get_habitat_checklist(distinct=True, dataset_id=checklist_id)
    elif page == 'species':
        names = get_species_checklist(distinct=True, dataset_id=checklist_id)
    else:
        raise NotImplementedError()
    name_and_code = names.filter_by(code=subject).first()[1]
    _, _, name = name_and_code.partition(subject)
    full = get_subject_refvals_mixed(page, subject)

    changed_dict = {}
    for (region, d1) in full.items():
        for (sheet_name, d2) in d1.items():
            if sheet_name not in changed_dict.keys():
                changed_dict[sheet_name] = {}
            changed_dict[sheet_name][region] = d2

    wb = Workbook()
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

    for (sheet_name, d1) in changed_dict.items():
        ws = wb.create_sheet()
        ws.title = REFGROUPS.get(sheet_name, sheet_name.title())
        operators = d1.values()[0].keys()
        ws.append(['COD SPECIE', 'NUME', 'BIOREGIUNE'] + operators)
        for (bioregion, d2) in d1.items():
            ws.append([subject, name, bioregion] +
                      [d2.get(operator) for operator in operators])

    response = Response(save_virtual_workbook(wb), mimetype=MIMETYPE)
    response.headers.add('Content-Disposition',
                         'attachment; filename={}.xlsx'.format(subject))
    return response


@aggregation.route('/admin/reference_values/<page>/download',
                   methods=['GET'])
def download_all_refvals(page):
    refvals = get_refvals(page)
    hd_table = LuHabitattypeCodes if page == 'habitat' else LuHdSpecies
    code_to_name = {str(lu.code): lu.display_name for lu in hd_table.query}

    data_struct = get_struct(refvals)

    wb = Workbook()
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))

    for group, fields in data_struct.iteritems():
        ws = wb.create_sheet()
        ws.title = group
        ws.append(fields)

    sorted_keys = sorted(refvals.keys())
    for key in sorted_keys:
        code, region = key.split('-')
        for group, values in refvals[key].iteritems():
            xls_group = REFGROUPS[group]
            ws = wb[xls_group]
            values = {k.upper(): v for k, v in values.iteritems()}
            values['COD'] = code
            values['BIOREGIUNE'] = region
            values['NUME'] = code_to_name.get(code, None)
            to_insert = [values.get(k, None) for k in data_struct[xls_group]]
            ws.append(to_insert)

    response = Response(save_virtual_workbook(wb), mimetype=MIMETYPE)
    response.headers.add('Content-Disposition',
                         'attachment; filename={}.xlsx'.format(page))
    return response


def parse_checklist_ref(checklist_qs):
    return {(item.code, item.bio_region): item for item in checklist_qs}
